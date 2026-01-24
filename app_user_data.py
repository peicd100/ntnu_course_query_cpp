from __future__ import annotations

import json
import os
import time
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app_constants import user_data_store_path
from app_utils import (
    format_cid4,
    parse_cid_to_int,
    sanitize_folder_name,
    truthy_flag,
)


def user_root_dir(course_excel_path: str) -> str:
    root = user_data_store_path()
    os.makedirs(root, exist_ok=True)
    return os.fspath(root)




def user_dir(course_excel_path: str, username: str) -> str:
    root = user_root_dir(course_excel_path)
    u = sanitize_folder_name(username)
    path = os.path.join(root, u)
    os.makedirs(path, exist_ok=True)
    return path


HISTORY_SUBDIR = "history"
BEST_SCHEDULE_SUBDIR = "best_schedule"
BEST_SCHEDULE_FILE_PREFIX = "學分數："


def _ensure_subdir(base_dir: str, subdir: str) -> str:
    if not base_dir:
        return ""
    target = os.path.join(base_dir, subdir)
    os.makedirs(target, exist_ok=True)
    return target


def history_dir_path(user_dir_path: str) -> str:
    return _ensure_subdir(user_dir_path, HISTORY_SUBDIR)


def best_schedule_dir_path(user_dir_path: str) -> str:
    return _ensure_subdir(user_dir_path, BEST_SCHEDULE_SUBDIR)


def unique_login_file_path(course_excel_path: str, username: str, ts: str) -> str:
    udir = user_dir(course_excel_path, username)
    history_dir = history_dir_path(udir)
    base = f"{ts}.xlsx"
    path = os.path.join(history_dir or udir, base)
    if not os.path.exists(path):
        return path
    k = 1
    while True:
        p2 = os.path.join(udir, f"{ts}_{k}.xlsx")
        if not os.path.exists(p2):
            return p2
        k += 1


def list_user_history_files(user_dir_path: str) -> List[str]:
    if not user_dir_path or not os.path.isdir(user_dir_path):
        return []
    files: List[str] = []
    seen: Set[str] = set()
    history_dir = history_dir_path(user_dir_path)
    if history_dir and os.path.isdir(history_dir):
        for fn in os.listdir(history_dir):
            if not fn.lower().endswith(".xlsx"):
                continue
            path = os.path.join(history_dir, fn)
            if not os.path.isfile(path):
                continue
            files.append(path)
            seen.add(os.path.abspath(path))
    for fn in os.listdir(user_dir_path):
        path = os.path.join(user_dir_path, fn)
        if not os.path.isfile(path):
            continue
        abs_path = os.path.abspath(path)
        if abs_path in seen:
            continue
        if not fn.lower().endswith(".xlsx"):
            continue
        if fn.startswith(BEST_SCHEDULE_FILE_PREFIX):
            continue
        files.append(path)
    files.sort(key=lambda p: os.path.basename(p), reverse=True)
    return files




def list_all_users(course_excel_path: str) -> List[str]:
    root = user_root_dir(course_excel_path)
    if not os.path.isdir(root):
        return []
    names: List[str] = []
    for entry in os.listdir(root):
        p = os.path.join(root, entry)
        if os.path.isdir(p):
            names.append(entry)
    names.sort(key=lambda s: s.lower())
    return names


BEST_SCHEDULE_CACHE_FILENAME = "best_schedule_cache.json"


def best_schedule_cache_path(user_dir_path: str) -> str:
    cache_dir = best_schedule_dir_path(user_dir_path)
    if not cache_dir:
        return ""
    return os.path.join(cache_dir, BEST_SCHEDULE_CACHE_FILENAME)


def load_best_schedule_cache(user_dir_path: str) -> Optional[Dict]:
    path = best_schedule_cache_path(user_dir_path)
    if not path or not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def save_best_schedule_cache(
    user_dir_path: str,
    favorites_ids: List[int],
    locked_ids: List[int],
    filenames: List[str],
) -> None:
    if not user_dir_path:
        return
    path = best_schedule_cache_path(user_dir_path)
    if not path:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    payload = {
        "favorites": [int(x) for x in favorites_ids],
        "locked": [int(x) for x in locked_ids],
        "files": list(filenames),
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_user_file(xlsx_path: str) -> Tuple[Set[int], Set[int], Dict[int, int], Set[int]]:
    """
    回傳：favorites, included, seq_map, locked_set
    相容舊檔：
      - 舊檔欄位：開課序號 / 課表 / 加入順序
      - 新檔欄位：開課序號 / 課表 / 鎖定 / 加入順序
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"找不到使用者檔案：{xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    if "我的最愛" not in wb.sheetnames:
        raise RuntimeError("使用者檔案缺少工作表「我的最愛」。")

    ws = wb["我的最愛"]
    fav: Set[int] = set()
    inc: Set[int] = set()
    locked: Set[int] = set()
    seq: Dict[int, int] = {}

    row_no = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        row_no += 1
        if not row or row[0] is None:
            break

        cid = parse_cid_to_int(row[0])
        if cid is None:
            continue
        cid_i = int(cid)
        fav.add(cid_i)

        flag_course = row[1] if len(row) >= 2 else 0

        lock_flag = False
        join_seq = None

        if len(row) >= 4:
            lock_flag = truthy_flag(row[2])
            join_seq = row[3]
        elif len(row) >= 3:
            join_seq = row[2]
        else:
            join_seq = None

        if truthy_flag(flag_course):
            inc.add(cid_i)
        if lock_flag:
            locked.add(cid_i)

        s = None
        if join_seq is not None:
            try:
                s = int(join_seq)
            except Exception:
                s = None
        if s is None:
            s = row_no
        seq[cid_i] = int(s)

    # 鎖定課程必須強制顯示於課表
    inc |= locked

    return fav, (inc & fav), seq, (locked & fav)


def save_user_file(
    xlsx_path: str,
    username: str,
    favorites_ids: Set[int],
    included_ids_sorted: np.ndarray,
    locked_ids_sorted: np.ndarray,
    fav_seq: Dict[int, int],
    courses_df,
) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    items = [(int(fav_seq.get(int(cid), 10**12)), int(cid)) for cid in favorites_ids]
    items.sort(key=lambda x: (x[0], x[1]))

    def _sorted_has(arr: np.ndarray, cid: int) -> bool:
        if arr is None or arr.size == 0:
            return False
        x = np.int64(int(cid))
        p = int(np.searchsorted(arr, x, side="left"))
        return 0 <= p < arr.size and int(arr[p]) == int(x)

    ws_fav = wb.create_sheet("我的最愛")
    ws_fav.append([f"使用者：{username}（此檔案由程式自動產生）"])
    ws_fav.append(["開課序號", "課表", "鎖定", "加入順序"])

    for s, cid in items:
        in_course = 1 if _sorted_has(included_ids_sorted, int(cid)) else 0
        is_lock = 1 if _sorted_has(locked_ids_sorted, int(cid)) else 0
        ws_fav.append([format_cid4(int(cid)), in_course, is_lock, int(s)])

    ws_fav.append([])
    ws_fav.append(["最後更新", time.strftime("%Y-%m-%d %H:%M:%S")])

    ws_tt = wb.create_sheet("課表匯出")
    out_cols = [c for c in courses_df.columns if not str(c).startswith("_")] + ["_tba", "_slots"]

    if included_ids_sorted.size:
        subset = courses_df[courses_df["_cid"].isin([int(x) for x in included_ids_sorted.tolist()])].copy()
    else:
        subset = courses_df.iloc[0:0].copy()

    if not subset.empty:
        subset["_slots"] = subset["_slots"].apply(lambda x: json.dumps(x, ensure_ascii=False))
        if "系所" in subset.columns and "中文課程名稱" in subset.columns:
            subset = subset.sort_values(["系所", "中文課程名稱", "_cid"], kind="mergesort")

    ws_tt.append([f"使用者：{username}（顯示於課表的課程匯出）"])
    ws_tt.append(out_cols)
    if not subset.empty:
        for r in dataframe_to_rows(subset[out_cols], index=False, header=False):
            ws_tt.append(r)

    wb.save(xlsx_path)
