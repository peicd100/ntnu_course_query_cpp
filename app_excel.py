from __future__ import annotations

import io
import os
import zipfile
from typing import List, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app_constants import COURSE_SHEET_CANDIDATES, REQUIRED_COLUMNS
from app_utils import (
    parse_cid_to_int,
    parse_time_text,
    slots_set_to_masks,
    format_cid4,
    _slot_sort_key,
    parse_gened_categories_from_course_name,
    strip_bracket_text_for_timetable,
)


class ExcelFormatError(RuntimeError):
    pass


def _patch_xlsx_namespaces_inplace(xlsx_path: str) -> bool:
    repls = {
        b"http://purl.oclc.org/ooxml/spreadsheetml/main": b"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        b"http://purl.oclc.org/ooxml/officeDocument/relationships": b"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        b"http://purl.oclc.org/ooxml/spreadsheetml/2009/9/main": b"http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    }

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        needs = False
        for name in zin.namelist():
            if name.endswith(".xml") and b"purl.oclc.org/ooxml" in zin.read(name):
                needs = True
                break

    if not needs:
        return False

    backup_path = xlsx_path + ".bak.xlsx"
    if not os.path.exists(backup_path):
        with open(xlsx_path, "rb") as fsrc, open(backup_path, "wb") as fdst:
            fdst.write(fsrc.read())

    buf = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml"):
                for a, b in repls.items():
                    data = data.replace(a, b)
            zout.writestr(item, data)

    with open(xlsx_path, "wb") as f:
        f.write(buf.getvalue())

    return True


def ensure_excel_readable(excel_path: str) -> str:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"找不到檔案：{excel_path}")

    ext = os.path.splitext(excel_path)[1].lower()

    if ext == ".xls":
        try:
            pd.ExcelFile(excel_path)
            return excel_path
        except Exception as e:
            raise ExcelFormatError(
                "讀取 .xls 需要安裝 xlrd。\n"
                "請執行：conda install -c conda-forge xlrd\n\n"
                f"原始錯誤：{e}"
            ) from e

    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            wb = load_workbook(excel_path, data_only=True)
            if not wb.sheetnames:
                raise ExcelFormatError("Workbook 沒有任何工作表。")
            return excel_path
        except Exception:
            patched = _patch_xlsx_namespaces_inplace(excel_path)
            if not patched:
                raise
            wb2 = load_workbook(excel_path, data_only=True)
            if not wb2.sheetnames:
                raise ExcelFormatError("已嘗試修補 Excel，但仍讀不到工作表。")
            return excel_path

    raise ExcelFormatError(f"不支援的 Excel 副檔名：{ext}（僅支援 .xls / .xlsx）")


def get_sheetnames(excel_path: str) -> List[str]:
    xls = pd.ExcelFile(excel_path)
    return list(xls.sheet_names)


def _pick_course_sheet(sheetnames: Sequence[str]) -> str:
    for s in COURSE_SHEET_CANDIDATES:
        if s in sheetnames:
            return s
    return sheetnames[0]


def _build_courses_df_from_raw(raw: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED_COLUMNS if c not in raw.columns]
    if missing:
        raise ExcelFormatError("Excel 欄位不足，缺少：" + ", ".join(missing))

    df = raw.copy()

    cid_series = df["開課序號"].apply(parse_cid_to_int)
    df = df[cid_series.notna()].copy()
    df["_cid"] = cid_series.loc[df.index].astype(int)
    df["開課序號"] = df["_cid"].apply(format_cid4)

    for col in ["開課代碼", "系所", "中文課程名稱", "教師", "必/選", "全/半", "地點時間"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    for col in ["學分", "限修人數", "選修人數"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Cache time parsing results to reduce repeated parsing on identical strings
    time_cache = {}
    parsed_results = []
    for text in df["地點時間"]:
        key = "" if text is None else str(text)
        res = time_cache.get(key)
        if res is None:
            res = parse_time_text(text)
            time_cache[key] = res
        parsed_results.append(res)
    df["_slots_set"] = [res.slots for res in parsed_results]
    df["_slots"] = [sorted(res.slots, key=_slot_sort_key) for res in parsed_results]
    df["_tba"] = [res.tba for res in parsed_results]

    masks = [slots_set_to_masks(slots) for slots in df["_slots_set"]]
    df["_mask_lo"] = [np.uint64(t[0]) for t in masks]
    df["_mask_hi"] = [np.uint64(t[1]) for t in masks]

    # Cache gened parsing to reduce repeated regex work
    gened_cache = {}
    gened_results = []
    for name in df["中文課程名稱"]:
        key = "" if name is None else str(name)
        res = gened_cache.get(key)
        if res is None:
            res = parse_gened_categories_from_course_name(name)
            gened_cache[key] = res
        gened_results.append(res)
    df["_gened_cats"] = gened_results

    cname_series = df["銝剜?隤脩??迂"].fillna("").astype(str)
    clean_name = cname_series.map(strip_bracket_text_for_timetable)
    serial_label = df["_cid"].map(format_cid4)
    df["_tt_label"] = (clean_name + "\n" + serial_label).str.strip()

    # Precompute lowercased fields for faster search (kept as internal columns)
    if "開課代碼" in df.columns:
        df["_code_lc"] = df["開課代碼"].str.lower()
    if "中文課程名稱" in df.columns:
        df["_cname_lc"] = df["中文課程名稱"].str.lower()
    if "教師" in df.columns:
        df["_teacher_lc"] = df["教師"].str.lower()
    if "系所" in df.columns:
        df["_dept_lc"] = df["系所"].str.lower()

    display_cols = [c for c in df.columns if not str(c).startswith("_")]
    
    # Using vectorized string concatenation is much faster than row-wise agg
    all_text_series = pd.Series([""] * len(df), index=df.index, dtype=str)
    for col in display_cols:
        all_text_series += df[col].astype(str).fillna("") + " "
    
    df["_alltext"] = all_text_series.str.lower()

    return df.reset_index(drop=True)


def load_courses_auto(excel_path: str) -> Tuple[pd.DataFrame, str]:
    sheetnames = get_sheetnames(excel_path)
    if not sheetnames:
        raise ExcelFormatError("Workbook 沒有任何工作表。")

    best_sheet = ""
    best_count = -1
    best_raw = None

    preferred = _pick_course_sheet(sheetnames)
    ordered = [preferred] + [s for s in sheetnames if s != preferred]

    for s in ordered:
        try:
            raw = pd.read_excel(excel_path, sheet_name=s)
            if any(c not in raw.columns for c in REQUIRED_COLUMNS):
                continue
            cnt = raw["開課序號"].apply(parse_cid_to_int).notna().sum()
            if cnt > best_count:
                best_count = int(cnt)
                best_sheet = s
                best_raw = raw
        except Exception:
            continue

    if best_raw is None:
        msg = "找不到包含必要欄位的工作表。\n\n"
        msg += "工作表清單：\n" + "\n".join([f"- {s}" for s in sheetnames])
        raise ExcelFormatError(msg)

    df = _build_courses_df_from_raw(best_raw)
    return df, best_sheet
